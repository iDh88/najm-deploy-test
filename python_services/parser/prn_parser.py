"""
PRN Bulk Upload Service
Accepts CSV or Excel file with employee PRN data.
Populates the prnWhitelist collection in Firestore.

Expected columns (flexible matching):
  PRN / Payroll Number / Employee ID
  First Name
  Last Name
  Rank (GD, PCA, BUT, CHF, SNF, YCA, CA, FO)
  Base (JED, RUH, DMM)
  Category (9Z, 782, etc.) — optional
  Seniority Date — optional
"""

from fastapi import APIRouter, HTTPException, BackgroundTasks, UploadFile, File, Form
from pydantic import BaseModel, Field
from typing import Optional
import openpyxl
import csv
import re
import io
import logging
from datetime import datetime

logger = logging.getLogger("cip.prn_parser")
router = APIRouter()

VALID_RANKS = {'GD', 'PCA', 'BUT', 'CHF', 'SNF', 'YCA', 'CA', 'FO'}
VALID_BASES = {'JED', 'RUH', 'DMM', 'MED'}

# Column name aliases for flexible matching
COLUMN_ALIASES = {
    'prn': ['prn', 'payroll', 'payroll number', 'employee id',
            'staff id', 'staff number', 'id', 'رقم الموظف'],
    'firstName': ['first name', 'firstname', 'first', 'الاسم الأول', 'fname'],
    'lastName': ['last name', 'lastname', 'last', 'اسم العائلة', 'lname', 'family name'],
    'rank': ['rank', 'position', 'الرتبة', 'grade'],
    'base': ['base', 'station', 'base station', 'القاعدة', 'الأساس'],
    'category': ['category', 'cat', 'الفئة', 'salary category'],
    'seniorityDate': ['seniority', 'seniority date', 'hire date', 'join date',
                      'تاريخ الأقدمية'],
}

# ─── Models ───────────────────────────────────────────────────────────────────

class PRNRecord(BaseModel):
    prn: str
    firstName: str
    lastName: str
    fullName: str
    rank: str
    base: str
    category: str = "9Z"
    seniorityDate: str = ""
    isValid: bool = True
    validationError: str = ""

class PRNUploadResult(BaseModel):
    success: bool
    totalRecords: int
    validRecords: int
    invalidRecords: int
    records: list[PRNRecord]
    errors: list[str]
    warnings: list[str]
    uploadedBy: str
    uploadedAt: str

# ─── Parser ───────────────────────────────────────────────────────────────────

class PRNParser:

    def __init__(self):
        self.errors: list[str] = []
        self.warnings: list[str] = []

    def parse(self, file_bytes: bytes, filename: str,
              uploaded_by: str) -> PRNUploadResult:

        ext = filename.lower().split('.')[-1]
        records = []

        try:
            if ext in ('xlsx', 'xls'):
                records = self._parse_excel(file_bytes)
            elif ext == 'csv':
                records = self._parse_csv(file_bytes)
            else:
                raise ValueError(f"Unsupported file type: .{ext}. Use .xlsx or .csv")
        except Exception as e:
            self.errors.append(str(e))

        valid = [r for r in records if r.isValid]
        invalid = [r for r in records if not r.isValid]

        if invalid:
            self.warnings.append(
                f"{len(invalid)} records skipped due to validation errors"
            )

        return PRNUploadResult(
            success=len(self.errors) == 0 and len(valid) > 0,
            totalRecords=len(records),
            validRecords=len(valid),
            invalidRecords=len(invalid),
            records=records,
            errors=self.errors,
            warnings=self.warnings,
            uploadedBy=uploaded_by,
            uploadedAt=datetime.utcnow().isoformat(),
        )

    def _parse_excel(self, file_bytes: bytes) -> list[PRNRecord]:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        # Find header row
        header_row, col_map = self._find_headers_excel(ws)
        if not col_map:
            raise ValueError(
                "Could not find required columns. "
                "Make sure your file has: PRN, First Name, Last Name, Rank, Base"
            )

        records = []
        for row_idx in range(header_row + 1, ws.max_row + 1):
            row = [
                ws.cell(row=row_idx, column=c).value
                for c in range(1, ws.max_column + 1)
            ]
            if all(v is None for v in row):
                continue
            record = self._parse_row(row, col_map)
            if record:
                records.append(record)

        return records

    def _parse_csv(self, file_bytes: bytes) -> list[PRNRecord]:
        # Try UTF-8 then Latin-1
        try:
            text = file_bytes.decode('utf-8-sig')
        except UnicodeDecodeError:
            text = file_bytes.decode('latin-1')

        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            raise ValueError("CSV file is empty")

        # Find header row
        col_map = self._find_headers_list(rows[0])
        if not col_map:
            raise ValueError(
                "Could not find required columns in CSV. "
                "Required: PRN, First Name, Last Name, Rank, Base"
            )

        records = []
        for row in rows[1:]:
            if not any(row):
                continue
            record = self._parse_row(row, col_map)
            if record:
                records.append(record)

        return records

    def _find_headers_excel(self, ws) -> tuple[int, dict]:
        for row_idx in range(1, min(10, ws.max_row + 1)):
            row = [
                str(ws.cell(row=row_idx, column=c).value or '').lower().strip()
                for c in range(1, ws.max_column + 1)
            ]
            col_map = self._find_headers_list(row)
            if col_map:
                return row_idx, col_map
        return 0, {}

    def _find_headers_list(self, header_row: list) -> dict:
        col_map = {}
        for field, aliases in COLUMN_ALIASES.items():
            for col_idx, cell_val in enumerate(header_row):
                cell_lower = str(cell_val).lower().strip()
                if any(alias in cell_lower for alias in aliases):
                    col_map[field] = col_idx
                    break
        # Must have at minimum PRN, rank, base
        required = {'prn', 'rank', 'base'}
        if not required.issubset(col_map.keys()):
            return {}
        return col_map

    def _parse_row(self, row: list, col_map: dict) -> Optional[PRNRecord]:
        def get(field: str, default='') -> str:
            idx = col_map.get(field)
            if idx is None or idx >= len(row):
                return default
            val = row[idx]
            return str(val).strip() if val is not None else default

        prn = re.sub(r'\D', '', get('prn')).strip()
        if not prn:
            return None

        # Name handling
        if 'firstName' in col_map and 'lastName' in col_map:
            first = get('firstName', '').title()
            last  = get('lastName', '').title()
            full  = f"{first} {last}".strip()
        elif 'firstName' in col_map:
            first = get('firstName', '').title()
            last  = ''
            full  = first
        else:
            # Try to parse full name from PRN column neighbour
            first = ''
            last  = ''
            full  = get('firstName', prn)

        rank = get('rank', '').upper().strip()
        base = get('base', '').upper().strip()[:3]
        category = get('category', '9Z').upper().strip() or '9Z'
        seniority = get('seniorityDate', '')

        # Validate
        error = ''
        is_valid = True

        if len(prn) < 4:
            error = f"PRN too short: '{prn}'"
            is_valid = False
        elif rank not in VALID_RANKS:
            # Try common mappings
            rank_map = {
                'ECONOMY CABIN ATTENDANT': 'YCA',
                'SENIOR CABIN ATTENDANT': 'SNF',
                'PREMIUM CABIN CREW': 'PCA',
                'GUEST DIRECTOR': 'GD',
                'BUTLER': 'BUT',
                'CHEF': 'CHF',
                'CAPTAIN': 'CA',
                'FIRST OFFICER': 'FO',
                'YCA': 'YCA', 'SNF': 'SNF', 'GD': 'GD',
                'PCA': 'PCA', 'BUT': 'BUT', 'CHF': 'CHF',
            }
            mapped = rank_map.get(rank.upper())
            if mapped:
                rank = mapped
            else:
                error = f"Unknown rank: '{rank}'"
                is_valid = False
        elif base not in VALID_BASES:
            error = f"Unknown base: '{base}'"
            is_valid = False

        return PRNRecord(
            prn=prn,
            firstName=first,
            lastName=last,
            fullName=full,
            rank=rank,
            base=base,
            category=category,
            seniorityDate=seniority,
            isValid=is_valid,
            validationError=error,
        )


# ─── API Endpoints ────────────────────────────────────────────────────────────

@router.post("/upload-prn-list", response_model=PRNUploadResult)
async def upload_prn_list(
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    uploaded_by: str = Form(...),
    replace_existing: bool = Form(False),
):
    """
    Upload a CSV or Excel file containing employee PRN data.
    Populates the prnWhitelist collection.
    Admin only.

    Required columns: PRN, First Name, Last Name, Rank, Base
    Optional columns: Category, Seniority Date
    """
    allowed = ('csv', 'xlsx', 'xls')
    ext = file.filename.lower().split('.')[-1]
    if ext not in allowed:
        raise HTTPException(
            status_code=422,
            detail=f"Unsupported format. Use: {', '.join(allowed)}"
        )

    file_bytes = await file.read()
    if len(file_bytes) > 20 * 1024 * 1024:
        raise HTTPException(status_code=422, detail="File too large. Maximum 20MB.")

    parser = PRNParser()
    result = parser.parse(
        file_bytes=file_bytes,
        filename=file.filename,
        uploaded_by=uploaded_by,
    )

    if result.validRecords > 0:
        background_tasks.add_task(
            _save_prn_to_firestore,
            records=result.records,
            uploaded_by=uploaded_by,
            replace_existing=replace_existing,
        )

    return result


@router.get("/prn-check/{prn}")
async def check_prn(prn: str):
    """
    Check if a PRN exists in the whitelist.
    Returns crew member info if found.
    Used by the signup flow.
    """
    try:
        from utils.firebase import get_firestore
        db = get_firestore()

        clean_prn = re.sub(r'\D', '', prn.strip())
        doc = db.collection("prnWhitelist").document(clean_prn).get()

        if doc.exists:
            data = doc.to_dict()
            return {
                "found": True,
                "prn": clean_prn,
                "fullName": data.get("fullName", ""),
                "firstName": data.get("firstName", ""),
                "lastName": data.get("lastName", ""),
                "rank": data.get("rank", ""),
                "base": data.get("base", ""),
                "category": data.get("category", "9Z"),
                "seniorityDate": data.get("seniorityDate", ""),
            }
        else:
            return {"found": False, "prn": clean_prn}

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/prn-stats")
async def get_prn_stats():
    """Return summary stats of the PRN whitelist."""
    try:
        from utils.firebase import get_firestore
        db = get_firestore()
        snap = db.collection("prnWhitelist").get()
        records = [d.to_dict() for d in snap]

        by_rank = {}
        by_base = {}
        for r in records:
            rank = r.get('rank', '?')
            base = r.get('base', '?')
            by_rank[rank] = by_rank.get(rank, 0) + 1
            by_base[base] = by_base.get(base, 0) + 1

        return {
            "total": len(records),
            "byRank": by_rank,
            "byBase": by_base,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/prn/{prn}")
async def delete_prn(prn: str, deleted_by: str):
    """Remove a PRN from the whitelist. Admin only."""
    try:
        from utils.firebase import get_firestore
        db = get_firestore()
        clean_prn = re.sub(r'\D', '', prn.strip())
        db.collection("prnWhitelist").document(clean_prn).delete()
        return {"success": True, "deleted": clean_prn}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# ─── Save to Firestore ────────────────────────────────────────────────────────

async def _save_prn_to_firestore(
    records: list[PRNRecord],
    uploaded_by: str,
    replace_existing: bool,
):
    try:
        from utils.firebase import get_firestore
        db = get_firestore()
        batch = db.batch()
        count = 0

        for record in records:
            if not record.isValid:
                continue

            ref = db.collection("prnWhitelist").document(record.prn)

            if not replace_existing:
                # Skip if already exists
                existing = ref.get()
                if existing.exists:
                    continue

            batch.set(ref, {
                "prn": record.prn,
                "firstName": record.firstName,
                "lastName": record.lastName,
                "fullName": record.fullName,
                "rank": record.rank,
                "base": record.base,
                "category": record.category,
                "seniorityDate": record.seniorityDate,
                "addedBy": uploaded_by,
                "addedAt": datetime.utcnow(),
                "isActive": True,
            }, merge=not replace_existing)

            count += 1
            if count % 400 == 0:
                batch.commit()
                batch = db.batch()

        if count % 400 != 0:
            batch.commit()

        # Save upload log
        db.collection("prnUploadLogs").add({
            "uploadedBy": uploaded_by,
            "uploadedAt": datetime.utcnow(),
            "totalSaved": count,
            "replaceExisting": replace_existing,
        })

        logger.info(f"Saved {count} PRN records to whitelist")

    except Exception as e:
        logger.error(f"PRN Firestore save failed: {e}", exc_info=True)
