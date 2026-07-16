"""
Excel Parser Service — Crew Intelligence Platform
Parses monthly Saudi Airlines roster Excel files.
Each sheet = one flight line (e.g. "Line 411").
"""

from fastapi import APIRouter, HTTPException, BackgroundTasks
from pydantic import BaseModel, Field
from typing import Optional
import openpyxl
import io
import re
import uuid
import logging
from datetime import datetime, timedelta
import pytz

logger = logging.getLogger("cip.parser")
router = APIRouter()

# ─── IATA → Timezone mapping ──────────────────────────────────────────────────
IATA_TIMEZONES = {
    "RUH": "Asia/Riyadh", "JED": "Asia/Riyadh", "DMM": "Asia/Riyadh",
    "MED": "Asia/Riyadh", "AHB": "Asia/Riyadh", "TIF": "Asia/Riyadh",
    "LHR": "Europe/London", "LGW": "Europe/London", "CDG": "Europe/Paris",
    "FRA": "Europe/Berlin", "AMS": "Europe/Amsterdam", "DXB": "Asia/Dubai",
    "AUH": "Asia/Dubai", "DOH": "Asia/Qatar", "KWI": "Asia/Kuwait",
    "BAH": "Asia/Bahrain", "CAI": "Africa/Cairo", "IST": "Europe/Istanbul",
    "BKK": "Asia/Bangkok", "SIN": "Asia/Singapore", "KUL": "Asia/Kuala_Lumpur",
    "CGK": "Asia/Jakarta", "DEL": "Asia/Kolkata", "BOM": "Asia/Kolkata",
    "LAX": "America/Los_Angeles", "JFK": "America/New_York", "ORD": "America/Chicago",
    "NRT": "Asia/Tokyo", "ICN": "Asia/Seoul", "PEK": "Asia/Shanghai",
    "SYD": "Australia/Sydney", "JNB": "Africa/Johannesburg",
    "GVA": "Europe/Zurich", "MAD": "Europe/Madrid", "FCO": "Europe/Rome",
    "MXP": "Europe/Rome", "MAN": "Europe/London", "BRU": "Europe/Brussels",
    "MUC": "Europe/Berlin", "VIE": "Europe/Vienna", "ZRH": "Europe/Zurich",
}

DOMESTIC_IATA = {"RUH", "JED", "DMM", "MED", "AHB", "TIF", "GIZ", "ELQ", "URY", "TUU", "AQI"}

COLUMN_ALIASES = {
    "flightNumber":  ["flight", "flt", "flt no", "flight number", "flight no", "رحلة", "رقم الرحلة"],
    "origin":        ["from", "dep", "origin", "departure", "المغادرة", "من"],
    "destination":   ["to", "arr", "dest", "destination", "arrival", "الوصول", "إلى"],
    "departureDate": ["date", "dep date", "departure date", "التاريخ"],
    "departureTime": ["dep time", "departure time", "std", "وقت المغادرة"],
    "arrivalTime":   ["arr time", "arrival time", "sta", "وقت الوصول"],
    "dutyStart":     ["duty start", "report", "report time", "بداية الواجب"],
    "dutyEnd":       ["duty end", "off duty", "release time", "نهاية الواجب"],
    "aircraftType":  ["aircraft", "ac type", "type", "a/c", "طائرة"],
    "payRate":       ["rate", "pay rate", "hourly rate", "معدل الأجر"],
}

# ─── Pydantic Models ──────────────────────────────────────────────────────────
class ParseRequest(BaseModel):
    userId: str
    month: str
    storageRef: str
    rank: str = ""   # GD | PCA | BUT | CHF | SNF | YCA | CA | FO

class ParsedLeg(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    lineId: str
    flightNumber: str
    origin: str
    destination: str
    legType: str
    departureLT: str
    arrivalLT: str
    departureUTC: str
    arrivalUTC: str
    dutyStart: str
    dutyEnd: str
    releaseTime: str
    blockHours: float = 0.0
    fdpHours: float = 0.0
    aircraftType: str = ""
    layover: bool = False
    layoverHours: float = 0.0
    payRate: float = 50.0
    estimatedPay: float = 0.0
    perDiem: float = 0.0
    legalityStatus: str = "legal"
    legalityFlags: list[str] = []
    restAfterHours: float = 0.0
    restBeforeHours: float = 0.0
    sequence: int = 0

class ParsedLine(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    lineNumber: str
    month: str
    userId: str
    rank: str = ""
    legs: list[ParsedLeg] = []
    summary: dict = {}
    destinations: list[str] = []
    daysOff: list[int] = []
    validationStatus: str = "valid"
    validationReport: dict = {}

class ParseResponse(BaseModel):
    success: bool
    linesProcessed: int
    lines: list[ParsedLine]
    errors: list[str]
    warnings: list[str]

# ─── Endpoints ────────────────────────────────────────────────────────────────
@router.post("/parse", response_model=ParseResponse)
async def parse_roster(request: ParseRequest, background_tasks: BackgroundTasks):
    try:
        excel_bytes = await _download_from_storage(request.storageRef)
        parser = RosterParser(user_id=request.userId, month=request.month)
        result = parser.parse(excel_bytes)
        background_tasks.add_task(_save_to_firestore, result)
        return result
    except Exception as e:
        logger.error(f"Parse failed: {e}", exc_info=True)
        raise HTTPException(status_code=422, detail=str(e))

# ─── Core Parser Class ────────────────────────────────────────────────────────
class RosterParser:
    def __init__(self, user_id: str, month: str):
        self.user_id = user_id
        self.month = month
        self.errors: list[str] = []
        self.warnings: list[str] = []

    def parse(self, excel_bytes: bytes) -> ParseResponse:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
        lines = []
        for sheet_name in wb.sheetnames:
            line_num = self._extract_line_number(sheet_name)
            if line_num is None:
                self.warnings.append(f"Skipped sheet '{sheet_name}'")
                continue
            try:
                line = self._parse_sheet(wb[sheet_name], line_num)
                if line:
                    lines.append(line)
            except Exception as e:
                self.errors.append(f"Line {line_num}: {e}")
        return ParseResponse(
            success=len(self.errors) == 0,
            linesProcessed=len(lines),
            lines=lines,
            errors=self.errors,
            warnings=self.warnings,
        )

    def _extract_line_number(self, sheet_name: str) -> Optional[str]:
        match = re.search(r'\d+', sheet_name.strip())
        return match.group() if match else None

    def _parse_sheet(self, ws, line_number: str) -> Optional[ParsedLine]:
        header_row_idx, col_map = self._find_headers(ws)
        if not col_map:
            self.errors.append(f"Line {line_number}: Cannot identify headers")
            return None

        line_id = str(uuid.uuid4())
        legs = []
        for row_idx in range(header_row_idx + 1, ws.max_row + 1):
            row = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
            if all(v is None for v in row):
                continue
            try:
                leg = self._parse_row(row, col_map, line_id, len(legs), line_number)
                if leg:
                    legs.append(leg)
            except Exception as e:
                self.warnings.append(f"Line {line_number} row {row_idx}: {e}")

        if not legs:
            return None

        legs = self._compute_rest_intervals(legs)
        legs = self._annotate_legality(legs)
        summary = self._compute_summary(legs)

        return ParsedLine(
            id=line_id,
            lineNumber=line_number,
            month=self.month,
            userId=self.user_id,
            legs=legs,
            summary=summary,
            destinations=list(set(l.destination for l in legs)),
            daysOff=self._compute_days_off(legs),
            validationStatus="warnings" if self.warnings else "valid",
            validationReport={"errors": self.errors, "warnings": self.warnings},
        )

    def _find_headers(self, ws) -> tuple[int, dict]:
        for row_idx in range(1, min(10, ws.max_row + 1)):
            row = [str(ws.cell(row=row_idx, column=c).value or '').lower().strip()
                   for c in range(1, ws.max_column + 1)]
            col_map = {}
            for field, aliases in COLUMN_ALIASES.items():
                for col_idx, cell_val in enumerate(row):
                    if any(alias in cell_val for alias in aliases):
                        col_map[field] = col_idx
                        break
            if len(col_map) >= 4:
                return row_idx, col_map
        return 0, {}

    def _parse_row(self, row: list, col_map: dict, line_id: str,
                   sequence: int, line_number: str) -> Optional[ParsedLeg]:
        def get(field, default=None):
            idx = col_map.get(field)
            return row[idx] if idx is not None and idx < len(row) else default

        flight_number = str(get("flightNumber", "")).strip()
        if not flight_number or flight_number.lower() in ['none', 'off', 'do', 'day off', '']:
            return None

        origin = str(get("origin", "RUH")).upper().strip()[:3]
        destination = str(get("destination", "JED")).upper().strip()[:3]
        aircraft_type = str(get("aircraftType", "")).strip()
        pay_rate = float(get("payRate", 50) or 50)

        dep_date = get("departureDate")
        dep_time = get("departureTime")
        arr_time = get("arrivalTime")
        duty_start_val = get("dutyStart")
        duty_end_val = get("dutyEnd")

        if not dep_date:
            return None

        dep_dt = self._parse_datetime(dep_date, dep_time or "00:00")
        if not dep_dt:
            return None

        arr_dt = self._parse_datetime(dep_date, arr_time) if arr_time else dep_dt + timedelta(hours=1)
        if arr_dt and arr_dt < dep_dt:
            arr_dt += timedelta(days=1)

        duty_start = (self._parse_datetime(dep_date, duty_start_val)
                      if duty_start_val else dep_dt - timedelta(hours=1))
        duty_end = (self._parse_datetime(dep_date, duty_end_val)
                    if duty_end_val else arr_dt + timedelta(minutes=30) if arr_dt else None)
        if duty_end and duty_start and duty_end < duty_start:
            duty_end += timedelta(days=1)

        # UTC conversion
        origin_tz = pytz.timezone(IATA_TIMEZONES.get(origin, "Asia/Riyadh"))
        dest_tz = pytz.timezone(IATA_TIMEZONES.get(destination, "Asia/Riyadh"))
        dep_utc = origin_tz.localize(dep_dt).astimezone(pytz.utc).replace(tzinfo=None)
        arr_utc = (dest_tz.localize(arr_dt).astimezone(pytz.utc).replace(tzinfo=None)
                   if arr_dt else dep_utc + timedelta(hours=1))

        block_hours = max(0, (arr_utc - dep_utc).total_seconds() / 3600)
        fdp_hours = max(0, (duty_end - duty_start).total_seconds() / 3600) if duty_end and duty_start else 0
        release_time = (duty_end + timedelta(minutes=30)) if duty_end else arr_utc + timedelta(hours=1)

        leg_type = "domestic" if origin in DOMESTIC_IATA and destination in DOMESTIC_IATA else "international"
        per_diem = 150.0 if leg_type == "international" else 50.0
        estimated_pay = round((block_hours * pay_rate) + per_diem, 2)

        duty_start_str = duty_start.isoformat() if duty_start else dep_dt.isoformat()
        duty_end_str = duty_end.isoformat() if duty_end else arr_dt.isoformat() if arr_dt else dep_dt.isoformat()

        return ParsedLeg(
            id=str(uuid.uuid4()),
            lineId=line_id,
            flightNumber=flight_number,
            origin=origin,
            destination=destination,
            legType=leg_type,
            departureLT=dep_dt.isoformat(),
            arrivalLT=arr_dt.isoformat() if arr_dt else dep_dt.isoformat(),
            departureUTC=dep_utc.isoformat(),
            arrivalUTC=arr_utc.isoformat(),
            dutyStart=duty_start_str,
            dutyEnd=duty_end_str,
            releaseTime=release_time.isoformat(),
            blockHours=round(block_hours, 2),
            fdpHours=round(fdp_hours, 2),
            aircraftType=aircraft_type,
            payRate=pay_rate,
            estimatedPay=estimated_pay,
            perDiem=per_diem,
            sequence=sequence,
        )

    def _parse_datetime(self, date_val, time_val) -> Optional[datetime]:
        try:
            if isinstance(date_val, datetime):
                base = date_val.date()
            else:
                date_str = str(date_val).strip()
                for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d %b %Y"]:
                    try:
                        base = datetime.strptime(date_str, fmt).date()
                        break
                    except ValueError:
                        continue
                else:
                    return None

            if isinstance(time_val, datetime):
                t = time_val.time()
            else:
                time_str = str(time_val).strip().replace(".", ":").replace(" ", "")
                for fmt in ["%H:%M", "%H:%M:%S", "%I:%M%p", "%H%M"]:
                    try:
                        t = datetime.strptime(time_str, fmt).time()
                        break
                    except ValueError:
                        continue
                else:
                    return None

            return datetime.combine(base, t)
        except Exception:
            return None

    def _compute_rest_intervals(self, legs: list[ParsedLeg]) -> list[ParsedLeg]:
        legs = list(legs)
        for i in range(1, len(legs)):
            prev_release = datetime.fromisoformat(legs[i - 1].releaseTime)
            curr_duty_start = datetime.fromisoformat(legs[i].dutyStart)
            rest_hours = max(0, (curr_duty_start - prev_release).total_seconds() / 3600)
            legs[i] = legs[i].model_copy(update={"restBeforeHours": round(rest_hours, 2)})
            legs[i - 1] = legs[i - 1].model_copy(update={
                "restAfterHours": round(rest_hours, 2),
                "layover": rest_hours >= 6,
                "layoverHours": round(rest_hours, 2) if rest_hours >= 6 else 0.0,
            })
        return legs

    def _annotate_legality(self, legs: list[ParsedLeg]) -> list[ParsedLeg]:
        annotated = []
        for leg in legs:
            violations, warnings_list = [], []
            min_rest = 15.0 if leg.legType == "international" else 14.0
            if 0 < leg.restBeforeHours < min_rest - 0.5:
                violations.append(f"GACA-REST-{leg.legType[:3].upper()}-001")
            elif 0 < leg.restBeforeHours < min_rest:
                warnings_list.append("GACA-REST-WARN-001")
            if leg.fdpHours > 14:
                violations.append("GACA-FDP-001")
            elif leg.fdpHours > 13:
                warnings_list.append("GACA-FDP-WARN-001")

            status = "violation" if violations else "warning" if warnings_list else "legal"
            annotated.append(leg.model_copy(update={
                "legalityStatus": status,
                "legalityFlags": violations + warnings_list,
            }))
        return annotated

    def _compute_summary(self, legs: list[ParsedLeg]) -> dict:
        total_block = sum(l.blockHours for l in legs)
        total_duty = sum(l.fdpHours for l in legs)
        total_pay = sum(l.estimatedPay for l in legs)
        intl = sum(1 for l in legs if l.legType == "international")
        dom = sum(1 for l in legs if l.legType == "domestic")
        layovers = sum(1 for l in legs if l.layover)
        rest_periods = [l.restAfterHours for l in legs if l.restAfterHours > 0]
        min_rest = min(rest_periods) if rest_periods else 0
        near_min = sum(1 for r in rest_periods if 0 < r < 15)
        rest_score = max(0, min(100, 100 - near_min * 15 - max(0, (14 - min_rest) * 5)))
        salary_score = min(100, (total_pay / 15000) * 100)
        duty_dates = set(datetime.fromisoformat(l.departureLT).date() for l in legs)
        return {
            "totalLegs": len(legs),
            "totalBlockHours": round(total_block, 2),
            "totalDutyHours": round(total_duty, 2),
            "totalDutyDays": len(duty_dates),
            "internationalLegs": intl,
            "domesticLegs": dom,
            "layoverCount": layovers,
            "estimatedSalaryMin": round(total_pay * 0.9, 2),
            "estimatedSalaryMax": round(total_pay * 1.1, 2),
            "salaryScore": round(salary_score, 1),
            "restQualityScore": round(rest_score, 1),
            "compositeScore": round((salary_score + rest_score) / 2, 1),
        }

    def _compute_days_off(self, legs: list[ParsedLeg]) -> list[int]:
        if not legs:
            return []
        duty_dates = set(datetime.fromisoformat(l.departureLT).date() for l in legs)
        first = min(duty_dates)
        last = max(duty_dates)
        days_off, current = [], first
        while current <= last:
            if current not in duty_dates:
                ksa_dow = (current.weekday() + 1) % 7  # Sun=0 in KSA
                days_off.append(ksa_dow)
            current += timedelta(days=1)
        return days_off

async def _download_from_storage(storage_ref: str) -> bytes:
    try:
        from utils.firebase import get_storage
        bucket = get_storage()
        blob = bucket.blob(storage_ref)
        return blob.download_as_bytes()
    except Exception as e:
        raise HTTPException(status_code=404, detail=f"File not found: {storage_ref}")

async def _save_to_firestore(result: ParseResponse):
    try:
        from utils.firebase import get_firestore
        from google.cloud.firestore import SERVER_TIMESTAMP
        db = get_firestore()
        for line in result.lines:
            line_ref = db.collection("flightLines").document(line.id)
            data = line.model_dump(exclude={"legs"})
            data["uploadedAt"] = datetime.utcnow()
            data["rank"] = line.rank
            line_ref.set(data)
            for leg in line.legs:
                line_ref.collection("legs").document(leg.id).set(leg.model_dump())
        logger.info(f"Saved {len(result.lines)} lines to Firestore")
    except Exception as e:
        logger.error(f"Firestore save failed: {e}", exc_info=True)
